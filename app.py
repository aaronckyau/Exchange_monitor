from flask import Flask, jsonify, render_template_string, request, send_file
import requests
from bs4 import BeautifulSoup
from datetime import datetime, timezone, timedelta

HKT = timezone(timedelta(hours=8))
import os
import io
import openpyxl
from dotenv import load_dotenv

load_dotenv()

app = Flask(__name__)

FMP_API_KEY = os.environ.get("FMP_API_KEY", "")

# ── HTML 模板 ────────────────────────────────────────────────
HTML = """<!DOCTYPE html>
<html lang="zh-TW">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1.0"/>
<title>USD/HKD 匯率監察</title>
<link rel="preconnect" href="https://fonts.googleapis.com"/>
<link href="https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&family=Noto+Sans+TC:wght@400;700&display=swap" rel="stylesheet"/>
<style>
  :root {
    --bg:        #f0f4f8;
    --panel:     #ffffff;
    --border:    #d1dbe8;
    --accent:    #0891b2;
    --accent2:   #2563eb;
    --accent3:   #d97706;
    --text:      #0f172a;
    --muted:     #64748b;
    --up:        #16a34a;
    --label-bg:  #e8eef5;
  }

  * { box-sizing: border-box; margin: 0; padding: 0; }

  body {
    background: var(--bg);
    color: var(--text);
    font-family: 'Noto Sans TC', sans-serif;
    min-height: 100vh;
    display: flex;
    flex-direction: column;
    align-items: center;
    padding: 40px 20px;
  }

  /* subtle grid bg */
  body::before {
    content: '';
    position: fixed; inset: 0;
    background-image:
      linear-gradient(var(--border) 1px, transparent 1px),
      linear-gradient(90deg, var(--border) 1px, transparent 1px);
    background-size: 40px 40px;
    opacity: 0.15;
    pointer-events: none;
    z-index: 0;
  }

  .wrapper { position: relative; z-index: 1; width: 100%; max-width: 860px; }

  header {
    display: flex;
    align-items: baseline;
    gap: 16px;
    margin-bottom: 36px;
  }

  header h1 {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 1.6rem;
    font-weight: 600;
    letter-spacing: 0.05em;
    color: var(--accent);
  }

  header .sub {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 0.75rem;
    color: var(--muted);
    letter-spacing: 0.1em;
  }

  /* ── Table ── */
  .rate-table {
    width: 100%;
    border-collapse: separate;
    border-spacing: 0;
    border: 1px solid var(--border);
    border-radius: 12px;
    overflow: hidden;
  }

  .rate-table th {
    background: var(--label-bg);
    font-family: 'IBM Plex Mono', monospace;
    font-size: 0.85rem;
    font-weight: 600;
    letter-spacing: 0.08em;
    padding: 18px 24px;
    text-align: center;
    border-bottom: 1px solid var(--border);
    color: var(--muted);
  }

  .rate-table th.boc  { color: var(--accent);  }
  .rate-table th.yfin { color: var(--accent2); }
  .rate-table th.fmp  { color: var(--accent3); }

  .rate-table th:not(:last-child),
  .rate-table td:not(:last-child) {
    border-right: 1px solid var(--border);
  }

  .rate-table td {
    background: var(--panel);
    padding: 16px 24px;
    text-align: center;
    font-family: 'IBM Plex Mono', monospace;
    font-size: 0.9rem;
    border-bottom: 1px solid var(--border);
    vertical-align: middle;
    transition: background 0.2s;
  }

  .rate-table tr:last-child td { border-bottom: none; }

  .rate-table td:first-child {
    text-align: left;
    color: var(--muted);
    font-size: 0.78rem;
    letter-spacing: 0.06em;
    background: var(--label-bg);
    white-space: nowrap;
  }

  .value {
    font-size: 1.25rem;
    font-weight: 600;
    color: var(--text);
    letter-spacing: 0.03em;
  }

  .value.mid  { color: var(--up); }
  .value.buy  { color: var(--accent2); }
  .value.sell { color: #f87171; }
  .value.dash { color: var(--muted); font-size: 1rem; }

  .sub-label {
    display: block;
    font-size: 0.68rem;
    color: var(--muted);
    margin-top: 4px;
    letter-spacing: 0.05em;
  }

  .ts {
    font-size: 0.72rem;
    color: var(--muted);
    letter-spacing: 0.04em;
  }

  /* ── Refresh btn ── */
  .btn-bar {
    margin-top: 28px;
    display: flex;
    align-items: center;
    gap: 16px;
  }

  .btn-refresh {
    display: flex;
    align-items: center;
    gap: 8px;
    padding: 12px 28px;
    background: transparent;
    border: 1px solid var(--accent);
    border-radius: 8px;
    color: var(--accent);
    font-family: 'IBM Plex Mono', monospace;
    font-size: 0.85rem;
    letter-spacing: 0.08em;
    cursor: pointer;
    transition: background 0.2s, color 0.2s;
  }

  .btn-refresh:hover {
    background: var(--accent);
    color: var(--bg);
  }

  .btn-refresh svg {
    transition: transform 0.5s;
  }

  .btn-refresh.loading svg {
    animation: spin 0.8s linear infinite;
  }

  @keyframes spin { to { transform: rotate(360deg); } }

  .last-refresh {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 0.75rem;
    color: var(--muted);
  }

  /* pulse on update */
  @keyframes pulse-green {
    0%   { background: rgba(8,145,178,0.12); }
    100% { background: var(--panel); }
  }
  .updated { animation: pulse-green 1s ease-out; }

  .btn-use {
    padding: 7px 18px;
    background: transparent;
    border: 1px solid var(--accent2);
    border-radius: 6px;
    color: var(--accent2);
    font-family: 'IBM Plex Mono', monospace;
    font-size: 0.75rem;
    letter-spacing: 0.06em;
    cursor: pointer;
    transition: background 0.2s, color 0.2s, opacity 0.2s;
  }
  .btn-use:hover:not(:disabled) {
    background: var(--accent2);
    color: #fff;
  }
  .btn-use:disabled {
    opacity: 0.35;
    cursor: not-allowed;
  }
  .btn-use.done {
    border-color: var(--up);
    color: var(--up);
  }

  .spinner-overlay {
    display: none;
    position: fixed; inset: 0;
    background: rgba(10,14,26,0.6);
    z-index: 99;
    align-items: center;
    justify-content: center;
    font-family: 'IBM Plex Mono', monospace;
    font-size: 1rem;
    color: var(--accent);
    letter-spacing: 0.1em;
  }
  .spinner-overlay.show { display: flex; }
</style>
</head>
<body>
<div class="wrapper">
  <header>
    <h1>USD / HKD</h1>
    <span class="sub">RATE MONITOR · 匯率監察</span>
  </header>

  <table class="rate-table" id="rateTable">
    <thead>
      <tr>
        <th></th>
        <th class="boc">BOC HK<br><span style="font-size:0.7rem;font-weight:400">中銀香港 電匯</span></th>
        <th class="yfin">Yahoo Finance<br><span style="font-size:0.7rem;font-weight:400">即時報價</span></th>
        <th class="fmp">FMP<br><span style="font-size:0.7rem;font-weight:400">Financial Modeling Prep</span></th>
      </tr>
    </thead>
    <tbody>
      <tr id="rowMid">
        <td>中間價 MID</td>
        <td><span id="bocMid" class="value mid">—</span><span class="sub-label">(買入+賣出)÷2</span></td>
        <td><span id="yfinPrice" class="value mid">—</span></td>
        <td><span id="fmpPrice" class="value mid">—</span></td>
      </tr>
      <tr id="rowBuy">
        <td>買入價 BUYING</td>
        <td><span id="bocBuy" class="value buy">—</span></td>
        <td><span class="value dash">—</span></td>
        <td><span class="value dash">—</span></td>
      </tr>
      <tr id="rowSell">
        <td>賣出價 SELLING</td>
        <td><span id="bocSell" class="value sell">—</span></td>
        <td><span class="value dash">—</span></td>
        <td><span class="value dash">—</span></td>
      </tr>
      <tr id="rowTs">
        <td>更新時間</td>
        <td><span id="bocTs" class="ts">—</span></td>
        <td><span id="yfinTs" class="ts">—</span></td>
        <td><span id="fmpTs" class="ts">—</span></td>
      </tr>
      <tr id="rowExport">
        <td style="color:var(--muted);font-size:0.78rem;letter-spacing:0.06em;">匯出 EXPORT</td>
        <td><button class="btn-use" id="useBoc" onclick="useRate('boc')" disabled>使用此匯率</button></td>
        <td><button class="btn-use" id="useYfin" onclick="useRate('yfin')" disabled>使用此匯率</button></td>
        <td><button class="btn-use" id="useFmp" onclick="useRate('fmp')" disabled>使用此匯率</button></td>
      </tr>
    </tbody>
  </table>

  <div class="btn-bar">
    <button class="btn-refresh" id="refreshBtn" onclick="fetchRates()">
      <svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round">
        <polyline points="23 4 23 10 17 10"/><polyline points="1 20 1 14 7 14"/>
        <path d="M3.51 9a9 9 0 0 1 14.85-3.36L23 10M1 14l4.64 4.36A9 9 0 0 0 20.49 15"/>
      </svg>
      REFRESH
    </button>
    <span class="last-refresh" id="lastRefresh">尚未載入</span>
  </div>
</div>

<div class="spinner-overlay" id="overlay">⟳ &nbsp;載入中...</div>

<script>
async function fetchRates() {
  const btn = document.getElementById('refreshBtn');
  const overlay = document.getElementById('overlay');
  btn.classList.add('loading');
  overlay.classList.add('show');

  try {
    const res = await fetch('/api/rates');
    const d = await res.json();

    // BOC
    const buy  = parseFloat(d.boc.buy);
    const sell = parseFloat(d.boc.sell);
    const mid  = ((buy + sell) / 2).toFixed(6);

    setText('bocMid',  d.boc.error  ? 'ERR' : mid);
    setText('bocBuy',  d.boc.error  ? 'ERR' : buy.toFixed(6));
    setText('bocSell', d.boc.error  ? 'ERR' : sell.toFixed(6));
    setText('bocTs',   d.boc.ts  || '—');

    // yfinance
    setText('yfinPrice', d.yfin.error ? 'ERR' : parseFloat(d.yfin.price).toFixed(4));
    setText('yfinTs',    d.yfin.ts || '—');

    // FMP
    setText('fmpPrice', d.fmp.error ? 'ERR' : parseFloat(d.fmp.price).toFixed(4));
    setText('fmpTs',    d.fmp.ts || '—');

    // pulse animation
    ['rowMid','rowBuy','rowSell','rowTs'].forEach(id => {
      const el = document.getElementById(id);
      el.classList.remove('updated');
      void el.offsetWidth;
      el.classList.add('updated');
    });

    document.getElementById('lastRefresh').textContent =
      '上次更新: ' + new Date().toLocaleTimeString('zh-HK');

    // 暫存資料供匯出用
    window._rateData = d;

    // 啟用匯出按鈕
    if (!d.boc.error)  { document.getElementById('useBoc').disabled  = false; }
    if (!d.yfin.error) { document.getElementById('useYfin').disabled = false; }
    if (!d.fmp.error)  { document.getElementById('useFmp').disabled  = false; }

  } catch(e) {
    document.getElementById('lastRefresh').textContent = '⚠ 載入失敗';
  }

  btn.classList.remove('loading');
  overlay.classList.remove('show');
}

function setText(id, val) {
  document.getElementById(id).textContent = val;
}

async function useRate(source) {
  const d = window._rateData;
  if (!d) return;

  let mid, ts;
  if (source === 'boc') {
    const buy = parseFloat(d.boc.buy);
    const sell = parseFloat(d.boc.sell);
    mid = parseFloat(((buy + sell) / 2).toFixed(6));
    ts  = d.boc.ts;
  } else if (source === 'yfin') {
    mid = parseFloat(parseFloat(d.yfin.price).toFixed(4));
    ts  = d.yfin.ts;
  } else {
    mid = parseFloat(parseFloat(d.fmp.price).toFixed(4));
    ts  = d.fmp.ts;
  }

  const btnId = source === 'boc' ? 'useBoc' : source === 'yfin' ? 'useYfin' : 'useFmp';
  const btn = document.getElementById(btnId);
  btn.textContent = '處理中...';
  btn.disabled = true;

  try {
    const res = await fetch('/api/download-rate', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ mid, ts })
    });
    if (!res.ok) throw new Error('server error');
    const blob = await res.blob();
    const url = URL.createObjectURL(blob);
    const a = document.createElement('a');
    a.href = url;
    a.download = 'exchange_rate_temp_hk.xlsx';
    a.click();
    URL.revokeObjectURL(url);
    btn.textContent = '✓ 已下載';
    btn.classList.add('done');
    setTimeout(() => {
      btn.textContent = '使用此匯率';
      btn.classList.remove('done');
      btn.disabled = false;
    }, 3000);
  } catch(e) {
    btn.textContent = '失敗';
    btn.disabled = false;
  }
}

// 自動載入
fetchRates();
</script>
</body>
</html>
"""

# ── BOC 抓取 ─────────────────────────────────────────────────
def get_boc_rate():
    url = "https://www.bochk.com/whk/rates/exchangeRatesHKD/exchangeRatesHKD-input.action?lang=en"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Referer": "https://www.bochk.com/en/investment/rates/hkdrates.html",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9,zh-TW;q=0.8",
    }
    try:
        resp = requests.get(url, headers=headers, timeout=15)
        resp.encoding = "UTF-8"
        soup = BeautifulSoup(resp.text, "html.parser")
        # 抓網頁更新時間
        boc_ts = None
        import re
        ts_tag = soup.find(string=re.compile(r"Information last updated at HK Time"))
        if ts_tag:
            m = re.search(r"(\d{4}/\d{2}/\d{2}\s+\d{2}:\d{2}:\d{2})", ts_tag)
            if m:
                boc_ts = m.group(1).replace("/", "-")

        for row in soup.find_all("tr"):
            cells = row.find_all(["td", "th"])
            texts = [c.get_text(strip=True) for c in cells]
            if any("USD" in t for t in texts):
                if len(texts) >= 3:
                    return {
                        "buy": texts[1],
                        "sell": texts[2],
                        "ts": boc_ts or datetime.now(HKT).strftime("%Y-%m-%d %H:%M:%S"),
                        "error": False,
                    }
        return {"error": True, "msg": "USD row not found"}
    except Exception as e:
        return {"error": True, "msg": str(e)}


# ── yfinance 抓取 ─────────────────────────────────────────────
def get_yfin_rate():
    try:
        import yfinance as yf
        ticker = yf.Ticker("USDHKD=X")
        price = ticker.fast_info["last_price"]
        return {
            "price": price,
            "ts": datetime.now(HKT).strftime("%Y-%m-%d %H:%M:%S"),
            "error": False,
        }
    except Exception as e:
        return {"error": True, "msg": str(e)}


# ── FMP 抓取 ──────────────────────────────────────────────────
def get_fmp_rate():
    try:
        url = f"https://financialmodelingprep.com/stable/quote?symbol=USDHKD&apikey={FMP_API_KEY}"
        resp = requests.get(url, timeout=10)
        data = resp.json()
        if data and isinstance(data, list):
            price = data[0].get("price")
            ts_unix = data[0].get("timestamp")
            ts = datetime.fromtimestamp(ts_unix, HKT).strftime("%Y-%m-%d %H:%M:%S") if ts_unix else datetime.now(HKT).strftime("%Y-%m-%d %H:%M:%S")
            return {
                "price": price,
                "ts": ts,
                "error": False,
            }
        return {"error": True, "msg": "No data"}
    except Exception as e:
        return {"error": True, "msg": str(e)}


# ── Routes ────────────────────────────────────────────────────
@app.route("/")
def index():
    return render_template_string(HTML)


@app.route("/api/rates")
def api_rates():
    return jsonify({
        "boc":  get_boc_rate(),
        "yfin": get_yfin_rate(),
        "fmp":  get_fmp_rate(),
    })


@app.route("/api/download-rate", methods=["POST"])
def download_rate():
    data = request.get_json()
    mid = data.get("mid")
    ts = data.get("ts")

    template_path = os.path.join(os.path.dirname(__file__), "exchange_rate_temp_hk.xlsx")
    wb = openpyxl.load_workbook(template_path)
    ws = wb["sheet1"]

    # 中間價寫入 D2, E2, F2, D3, E3, F3
    for cell in ["D2", "E2", "F2", "D3", "E3", "F3"]:
        ws[cell] = mid

    # 發布日期寫入 G2, G3（只存日期）
    try:
        ts_dt = datetime.strptime(ts, "%Y-%m-%d %H:%M:%S").date()
    except Exception:
        ts_dt = ts
    ws["G2"] = ts_dt
    ws["G3"] = ts_dt
    ws["G2"].number_format = "dd/mm/yyyy"
    ws["G3"].number_format = "dd/mm/yyyy"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        as_attachment=True,
        download_name="exchange_rate_temp_hk.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    app.run(debug=False, host="0.0.0.0", port=5050)